from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import TypeVar

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel

from nianlun.evaluation import EvaluationCase, RagEvaluator
from nianlun.evaluation.contracts import JudgeMetadata
from nianlun.evaluation.reporting.excel import HEADERS, export_excel

T = TypeVar("T", bound=BaseModel)


class UnusedJudge:
    metadata = JudgeMetadata(provider="fake", model="fake", temperature=0.0)

    async def generate_structured_output(self, *, prompt: str, schema: type[T]) -> T:
        raise AssertionError(
            f"empty answers must not invoke judge: {prompt!r} {schema}"
        )


class FailingJudge:
    metadata = JudgeMetadata(provider="fake", model="fake", temperature=0.0)

    async def generate_structured_output(self, *, prompt: str, schema: type[T]) -> T:
        del prompt, schema
        raise RuntimeError("judge failed")


def _case_payload() -> dict[str, object]:
    return {
        "question": "What is the answer?",
        "reference_answer": "The reference answer.",
        "actual_answer": "",
        "retrieval_contexts": [],
    }


def test_export_excel_writes_concise_latest_records(tmp_path: Path) -> None:
    input_path = tmp_path / "input.jsonl"
    results_path = tmp_path / "results.jsonl"
    output_path = tmp_path / "report.xlsx"
    case_payload = _case_payload()
    failed_case_payload = {**case_payload, "actual_answer": "An answer."}
    input_path.write_text(
        "\n".join(
            (
                json.dumps(case_payload),
                json.dumps({"not": "a valid case"}),
                json.dumps(case_payload),
                json.dumps(failed_case_payload),
            )
        )
        + "\n",
        encoding="utf-8",
    )
    outcome = asyncio.run(
        RagEvaluator(judge=UnusedJudge()).evaluate(
            EvaluationCase.model_validate(case_payload)
        )
    )
    failed_outcome = asyncio.run(
        RagEvaluator(judge=FailingJudge()).evaluate(
            EvaluationCase.model_validate(failed_case_payload)
        )
    )
    results_path.write_text(
        "\n".join(
            (
                json.dumps(
                    {
                        "input_line": 1,
                        "error": {
                            "code": "evaluation_task_failed",
                            "message": "failed",
                        },
                    }
                ),
                json.dumps(
                    {
                        "input_line": 1,
                        "result": outcome.model_dump(mode="json"),
                    }
                ),
                json.dumps(
                    {
                        "input_line": 2,
                        "error": {"code": "invalid_input", "message": "invalid"},
                    }
                ),
                json.dumps(
                    {
                        "input_line": 3,
                        "skipped": {
                            "code": "duplicate_case",
                            "duplicate_of_input_line": 1,
                        },
                    }
                ),
                json.dumps(
                    {
                        "input_line": 4,
                        "result": failed_outcome.model_dump(mode="json"),
                    }
                ),
            )
        )
        + "\n",
        encoding="utf-8",
    )

    export_excel(
        input_path=input_path,
        results_path=results_path,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    worksheet = workbook["评估结果"]
    assert tuple(cell.value for cell in worksheet[1]) == HEADERS
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:J4"
    assert worksheet.max_row == 4
    assert [cell.value for cell in worksheet[2]] == [
        1,
        "What is the answer?",
        "The reference answer.",
        None,
        "incorrect",
        "actual_answer is empty",
        "generation_empty",
        "actual_answer is empty",
        "completed",
        None,
    ]
    assert [cell.value for cell in worksheet[3]] == [
        2,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "failed",
        "invalid_input",
    ]
    assert [cell.value for cell in worksheet[4]] == [
        4,
        "What is the answer?",
        "The reference answer.",
        "An answer.",
        None,
        None,
        None,
        None,
        "failed",
        failed_outcome.error.code if failed_outcome.error else None,
    ]


def test_export_excel_rejects_source_paths(tmp_path: Path) -> None:
    input_path = tmp_path / "input.jsonl"
    results_path = tmp_path / "results.jsonl"
    input_path.write_text("", encoding="utf-8")
    results_path.write_text("", encoding="utf-8")

    with pytest.raises(ValueError, match="must differ"):
        export_excel(
            input_path=input_path,
            results_path=results_path,
            output_path=results_path,
        )
