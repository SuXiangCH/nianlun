"""Concise, human-readable Excel export for JSONL evaluation results."""

from __future__ import annotations

import json
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nianlun.evaluation.batch import CaseAdapter
from nianlun.evaluation.contracts.case import EvaluationCase
from nianlun.evaluation.contracts.outcome import EvaluationOutcome

HEADERS = (
    "input_line",
    "question",
    "reference_answer",
    "actual_answer",
    "verdict",
    "verdict_reason",
    "attribution",
    "attribution_reason",
    "status",
    "error_code",
)

_COLUMN_WIDTHS = (12, 36, 48, 48, 20, 48, 26, 48, 14, 28)
_FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True, slots=True)
class _ExportRow:
    input_line: int
    question: str = ""
    reference_answer: str = ""
    actual_answer: str = ""
    verdict: str = ""
    verdict_reason: str = ""
    attribution: str = ""
    attribution_reason: str = ""
    status: str = "failed"
    error_code: str = ""

    def values(self) -> tuple[int | str, ...]:
        return (
            self.input_line,
            self.question,
            self.reference_answer,
            self.actual_answer,
            self.verdict,
            self.verdict_reason,
            self.attribution,
            self.attribution_reason,
            self.status,
            self.error_code,
        )


def export_excel(
    *,
    input_path: Path,
    results_path: Path,
    output_path: Path,
    adapter: CaseAdapter | None = None,
) -> None:
    """Export the latest actionable JSONL result for each input line to Excel.

    The worksheet intentionally excludes nested evidence, routing, retry, and usage
    fields. Duplicate-case ``skipped`` records are omitted because their original
    input line already owns the outcome.
    """
    if output_path.resolve() in {input_path.resolve(), results_path.resolve()}:
        raise ValueError(
            "Excel output path must differ from input and JSONL result paths"
        )
    input_cases = _read_input_cases(input_path, adapter)
    rows = _read_latest_rows(results_path, input_cases)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "评估结果"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(tuple(_safe_cell_value(value) for value in row.values()))

    _format_worksheet(worksheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _read_input_cases(
    input_path: Path, adapter: CaseAdapter | None
) -> dict[int, EvaluationCase]:
    cases: dict[int, EvaluationCase] = {}
    with input_path.open(encoding="utf-8") as source:
        for input_line, line in enumerate(source, start=1):
            if not line.strip():
                continue
            try:
                raw = json.loads(line)
                if not isinstance(raw, Mapping):
                    continue
                cases[input_line] = (
                    adapter(raw) if adapter else EvaluationCase.model_validate(raw)
                )
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
    return cases


def _read_latest_rows(
    results_path: Path, input_cases: Mapping[int, EvaluationCase]
) -> list[_ExportRow]:
    latest: dict[int, _ExportRow] = {}
    with results_path.open(encoding="utf-8") as source:
        for line in source:
            try:
                payload = json.loads(line)
                if not isinstance(payload, Mapping):
                    continue
                input_line = _input_line(payload)
                row = _row_from_payload(
                    payload, input_line, input_cases.get(input_line)
                )
            except (KeyError, TypeError, ValueError, json.JSONDecodeError):
                continue
            if row is not None:
                latest[input_line] = row
    return [latest[input_line] for input_line in sorted(latest)]


def _input_line(payload: Mapping[str, Any]) -> int:
    input_line = payload["input_line"]
    if (
        isinstance(input_line, bool)
        or not isinstance(input_line, int)
        or input_line < 1
    ):
        raise ValueError("input_line must be a positive integer")
    return input_line


def _row_from_payload(
    payload: Mapping[str, Any], input_line: int, case: EvaluationCase | None
) -> _ExportRow | None:
    if "result" in payload:
        outcome = EvaluationOutcome.model_validate(payload["result"])
        return _row_from_outcome(input_line, case, outcome)
    if "error" in payload:
        error = payload["error"]
        if not isinstance(error, Mapping):
            raise ValueError("error must be an object")
        code = error.get("code")
        if not isinstance(code, str) or not code:
            raise ValueError("error.code must be a non-empty string")
        return _base_row(input_line, case, status="failed", error_code=code)
    if "skipped" in payload:
        return None
    raise ValueError("record must contain result, error, or skipped")


def _row_from_outcome(
    input_line: int, case: EvaluationCase | None, outcome: EvaluationOutcome
) -> _ExportRow:
    if outcome.evaluation_status.value == "failed":
        error_code = outcome.error.code if outcome.error is not None else ""
        return _base_row(input_line, case, status="failed", error_code=error_code)

    correctness = outcome.correctness
    if correctness is None:
        raise ValueError("completed outcome requires correctness")
    attribution = outcome.attribution
    return _base_row(
        input_line,
        case,
        verdict=correctness.value.value,
        verdict_reason=correctness.reason,
        attribution=attribution.value.value if attribution else "",
        attribution_reason=attribution.reason if attribution else "",
        status="completed",
    )


def _base_row(
    input_line: int,
    case: EvaluationCase | None,
    *,
    verdict: str = "",
    verdict_reason: str = "",
    attribution: str = "",
    attribution_reason: str = "",
    status: str,
    error_code: str = "",
) -> _ExportRow:
    return _ExportRow(
        input_line=input_line,
        question=case.question if case else "",
        reference_answer=case.reference_answer if case else "",
        actual_answer=case.actual_answer if case else "",
        verdict=verdict,
        verdict_reason=verdict_reason,
        attribution=attribution,
        attribution_reason=attribution_reason,
        status=status,
        error_code=error_code,
    )


def _safe_cell_value(value: int | str) -> int | str:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _format_worksheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column, width in enumerate(_COLUMN_WIDTHS, start=1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 22
